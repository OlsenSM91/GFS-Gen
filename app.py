import openpyxl
from openpyxl.styles import PatternFill
import subprocess
import sys
import os
import threading
from appgui import open_gui
from tkinter import filedialog
from datetime import datetime

def resource_path(relative_path):
    """Get absolute path to resource for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)

def is_valid_hour(value):
    """Helper function to validate hour values"""
    try:
        hour = int(str(value))
        return 1 <= hour <= 12
    except (ValueError, TypeError):
        return False

def highlight_cells(sheet, start_hour, frequency, row, is_medication=False, has_merged_pair=False):
    # Validate inputs before processing
    if not start_hour or not frequency:
        print(f"Missing start_hour or frequency for row {row}")
        return
        
    try:
        start_hour = int(start_hour)
        frequency = int(frequency)
        if not (1 <= start_hour <= 12):
            print(f"Invalid start hour {start_hour} for row {row}")
            return
    except (ValueError, TypeError):
        print(f"Invalid start_hour ({start_hour}) or frequency ({frequency}) for row {row}")
        return

    # Map of merged column pairs for treatments that have merged cells
    merged_columns = {'F', 'J', 'N', 'R', 'V', 'Z'}
    
    # Map hours to columns based on row 5 time labels
    time_columns = []
    hour_mapping = {}
    
    # Scan row 5 from D to AA for time values
    all_columns = [chr(i) for i in range(ord('D'), ord('Z') + 1)] + ['AA']
    for col in all_columns:
        cell = sheet[f"{col}5"]
        cell_value = cell.value

        try:
            if cell_value and is_valid_hour(cell_value):
                hour = int(str(cell_value))
                if col not in hour_mapping.values():
                    hour_mapping[hour] = col
                    time_columns.append(col)
        except (ValueError, TypeError):
            continue

    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Find all occurrences of start hour
    start_indices = []
    for i, col in enumerate(time_columns):
        cell_value = sheet[f"{col}5"].value
        if cell_value and int(str(cell_value)) == start_hour:
            start_indices.append(i)

    # Highlight cells for each starting position
    for start_idx in start_indices:
        current_idx = start_idx
        while current_idx < len(time_columns):
            col = time_columns[current_idx]
            
            # Highlight the main cell
            cell = f"{col}{row}"
            sheet[cell].fill = highlight_fill
            
            # Handle merged cells
            if has_merged_pair:
                if col in merged_columns:
                    # For columns with merged cells
                    cell_below = f"{col}{row + 1}"
                    sheet[cell_below].fill = highlight_fill
                else:
                    # For non-merged columns, still highlight both cells
                    cell_below = f"{col}{row + 1}"
                    sheet[cell_below].fill = highlight_fill
            elif is_medication:
                # Regular medication handling
                cell_below = f"{col}{row + 1}"
                sheet[cell_below].fill = highlight_fill
            
            # Move forward by frequency
            current_idx += frequency
            if current_idx >= len(time_columns):
                break

def fill_template(data):
    try:
        print("Starting to process the template...")
        template_path = resource_path("template.xlsx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file '{template_path}' not found.")

        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook.active

        # Set up fill colors for CPR/DNR
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Process medications first to ensure proper placeholder handling
        medications = data.get("medications", [])
        med_replacements = {}
        for i, medication in enumerate(medications, start=1):
            if i <= 8:
                name = medication.get("name", "").strip()
                dosage = medication.get("dosage", "").strip()
                combined_med = f"{name} {dosage}".strip()
                med_replacements[f"{{med{i}}}"] = combined_med

        # Basic replacements dictionary
        replacements = {
            "{cpr_dnr}": data.get("cpr_dnr", ""),
            "{patient}": data.get("patient", ""),
            "{chartnum}": data.get("chartnum", ""),
            "{date}": data.get("date", ""),
            "{owner}": data.get("owner", ""),
            "{problem}": data.get("problem", ""),
            "{dvm}": data.get("dvm", ""),
            "{age}": data.get("age", ""),
            "{sex}": data.get("sex", ""),
            "{ivcinfo}": data.get("ivcinfo", ""),
            "{techs}": data.get("techs", ""),
            "{initials}": data.get("initials", "")
        }

        # Add procedure replacements
        procedures = data.get("procedures", [])
        for i in range(6):  # Handle all 6 possible procedures
            num = i + 2
            date_placeholder = f"{{date{num}}}"
            note_placeholder = f"{{noted{num}}}"
            
            if i < len(procedures):
                proc = procedures[i]
                replacements[date_placeholder] = proc.get("date", "")
                replacements[note_placeholder] = proc.get("note", "")
            else:
                replacements[date_placeholder] = ""
                replacements[note_placeholder] = ""

        # Handle all replacements
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_value = cell.value
                    
                    # Handle CPR/DNR with highlighting
                    if cell_value == "{cpr_dnr}":
                        cpr_dnr_value = data.get("cpr_dnr", "")
                        cell.value = cpr_dnr_value
                        if cpr_dnr_value == "CPR":
                            cell.fill = yellow_fill
                        elif cpr_dnr_value == "DNR":
                            cell.fill = red_fill
                    # Handle medication replacements
                    elif cell_value in med_replacements:
                        cell.value = med_replacements[cell_value]
                    # Handle inline variables
                    elif "{a}" in cell_value:
                        cell.value = cell_value.replace("{a}", data.get("a", ""))
                    elif "{e}" in cell_value:
                        cell.value = cell_value.replace("{e}", data.get("e", ""))
                    # Handle other placeholders
                    elif cell_value in replacements:
                        cell.value = replacements[cell_value]
                    # Clear any remaining placeholders
                    elif cell_value.startswith("{") and cell_value.endswith("}"):
                        cell.value = ""

        # Update treatment row mappings with merged cell information
        treatment_rows = {
            "Temperature": {"row": 6, "merged": False},
            "Pulse": {"row": 7, "merged": False},
            "Respiratory Rate": {"row": 8, "merged": False},
            "MMCRT": {"row": 9, "merged": True},
            "IV Fluids / Rate": {"row": 11, "merged": True},
            "Additives": {"row": 13, "merged": True},
            "Check IVC": {"row": 15, "merged": False},
            "Walk/Litter": {"row": 16, "merged": False},
            "Urine (+/-)": {"row": 17, "merged": False},
            "Stool (+/-)": {"row": 18, "merged": False},
            "Vomit (+/-)": {"row": 19, "merged": False},
            "Food / Water": {"row": 20, "merged": True}
        }

        for treatment_name, info in treatment_rows.items():
            if treatment_name in data.get("treatments", {}):
                details = data["treatments"][treatment_name]
                start_hour = details.get("start_hour", "")
                frequency = details.get("frequency", "")
                
                if start_hour and frequency and start_hour != "Start Hour" and frequency != "Frequency":
                    try:
                        highlight_cells(sheet, start_hour, frequency, info["row"], has_merged_pair=info["merged"])
                    except Exception as e:
                        print(f"Error highlighting treatment {treatment_name}: {str(e)}")

        # Process medication highlighting
        for i, medication in enumerate(medications, start=1):
            if i <= 8:
                row_num = 21 + (i * 2 - 1)  # Calculate row dynamically
                start_hour = medication.get("start_hour", "")
                frequency = medication.get("frequency", "")
                
                if start_hour and frequency and start_hour != "Start Hour" and frequency != "Frequency":
                    try:
                        highlight_cells(sheet, start_hour, frequency, row_num, is_medication=True)
                    except Exception as e:
                        print(f"Error highlighting medication {i}: {str(e)}")

        # Generate default filename
        patient_name = data.get("patient", "").strip()
        if patient_name.startswith("Enter "): patient_name = ""
        today = datetime.now().strftime("%Y.%m.%d")
        default_filename = f"{today}-{patient_name}-Flow-Chart.xlsx"
        
        # Get user's documents directory
        documents_path = os.path.expanduser("~/Documents")
        
        # Open save dialog
        output_path = filedialog.asksaveasfilename(
            initialdir=documents_path,
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if not output_path:
            print("Save cancelled by user")
            return

        # Save the updated sheet
        workbook.save(output_path)
        print("Template processing completed. Saving the file.")

        # Open the file with the default application
        if os.name == "nt":  # Windows
            os.startfile(output_path)
        elif os.name == "posix":  # macOS or Linux
            subprocess.call(["open" if "darwin" in os.sys.platform else "xdg-open", output_path])

        print("Flow sheet generated successfully and opened.")

    except Exception as e:
        print(f"Error generating flow sheet: {str(e)}")
        raise

def start_processing(data):
    try:
        fill_template(data)
    except Exception as e:
        print(f"Error in processing thread: {str(e)}")

if __name__ == "__main__":
    def handle_submit(data):
        print("Data received from GUI:", data)
        thread = threading.Thread(target=start_processing, args=(data,))
        thread.daemon = True
        thread.start()

    open_gui(handle_submit)